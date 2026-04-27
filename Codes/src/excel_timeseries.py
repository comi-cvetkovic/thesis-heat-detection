"""Utilities for reading timestamped Excel sheets with offset headers."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


def find_header_row(
    workbook_path: str | Path,
    sheet_name: str,
    first_column_name: str = "Time stamp",
    max_scan_rows: int = 100,
) -> int:
    """Return the zero-based row index containing the sheet header."""
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name]

    try:
        for row_number, row in enumerate(
            worksheet.iter_rows(min_row=1, max_row=max_scan_rows, values_only=True),
            start=1,
        ):
            first_value = row[0] if row else None
            if isinstance(first_value, str) and first_value.strip().lower() == first_column_name.lower():
                return row_number - 1
    finally:
        workbook.close()

    raise ValueError(
        f"Could not find header row starting with {first_column_name!r} "
        f"in sheet {sheet_name!r}."
    )


def read_timeseries_sheet(
    workbook_path: str | Path,
    sheet_name: str,
    timestamp_col: str = "Time stamp",
    dayfirst: bool = True,
) -> pd.DataFrame:
    """Read an Excel sheet whose real header row is preceded by notes/blanks."""
    header_row = find_header_row(workbook_path, sheet_name, first_column_name=timestamp_col)
    frame = pd.read_excel(workbook_path, sheet_name=sheet_name, header=header_row)
    frame = frame.dropna(how="all")

    if timestamp_col not in frame.columns:
        raise ValueError(f"Expected timestamp column {timestamp_col!r} in sheet {sheet_name!r}.")

    frame[timestamp_col] = pd.to_datetime(frame[timestamp_col], errors="coerce", dayfirst=dayfirst)
    frame = frame.dropna(subset=[timestamp_col])
    frame = frame.sort_values(timestamp_col).reset_index(drop=True)
    return frame
